"""
inference.py

Evaluates a trained LoRA adapter on the flower classification test set.

For each test image, the base VLM (with the adapter applied) is prompted
to identify the flower species, its JSON output is parsed and checked
against the ground-truth label, and the result is written to an Excel
report alongside a thumbnail of the image. A final accuracy summary is
printed once all examples have been scored.

An eval_config.json capturing the exact settings used is saved next to
the report, so a given report can always be traced back to how it was
produced.
"""

import argparse
import json
import os
import re

import openpyxl
import torch
from openpyxl.drawing.image import Image as OpenPyxlImage
from peft import PeftModel
from tqdm import tqdm
from transformers import AutoModelForImageTextToText, AutoProcessor, set_seed

from Config import model_config, build_bnb_config
from data_format import dataset_format, DEFAULT_DATASET

EVAL_PROMPT = (
    "You are an expert botanist, Identify the flower species in this image "
    "and output its label ID number in valid JSON."
)


def load_run_config_defaults(adapter_path: str) -> dict:
    """
    If the adapter directory has a run_config.json (written by the current
    train.py), pull dataset/seed/model_id from it so eval defaults to
    exactly what the checkpoint was trained with, unless the user
    explicitly overrides a flag on the CLI.
    """
    path = os.path.join(adapter_path, "run_config.json")
    if not os.path.exists(path):
        return {}
    with open(path) as f:
        run_config = json.load(f)
    return {
        "dataset": run_config.get("dataset"),
        "base_model_id": run_config.get("model_id"),
        "seed": run_config.get("seed"),
    }


def parse_args():
    # First pass: just get adapter_path so we can seed defaults from its
    # saved run_config.json (if any) before building the full parser.
    pre = argparse.ArgumentParser(add_help=False)
    pre.add_argument("--adapter_path", type=str, default="finalmodel/")
    pre_args, _ = pre.parse_known_args()
    inherited = load_run_config_defaults(pre_args.adapter_path)

    p = argparse.ArgumentParser(
        description="Run flower-classification eval for a trained adapter and write an Excel report.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    p.add_argument("--adapter_path", type=str, default=pre_args.adapter_path,
                    help="Path to the trained LoRA adapter directory.")
    p.add_argument("--base_model_id", type=str, default=inherited.get("base_model_id") or "HuggingFaceTB/SmolVLM2-2.2B-Instruct",
                    help="Base vision-language model the adapter was trained on top of.")
    p.add_argument("--dataset", type=str, default=inherited.get("dataset") or DEFAULT_DATASET,
                    help="HuggingFace dataset id to evaluate on.")
    p.add_argument("--prompt", type=str, default=EVAL_PROMPT,
                    help="Prompt used at eval time (can differ from the training prompt).")
    p.add_argument(
        "--seed",
        type=int,
        default=inherited.get("seed") if inherited.get("seed") is not None else 42,
        required=False,
        help="Random seed for dataset shuffling and generation.",
    )

    p.add_argument("--load_in_4bit", action=argparse.BooleanOptionalAction, default=True,
                    help="Load the base model in 4-bit precision via bitsandbytes.")
    p.add_argument("--bnb_quant_type", type=str, default="nf4",
                    help="bitsandbytes 4-bit quantization type.")
    p.add_argument("--bnb_double_quant", action=argparse.BooleanOptionalAction, default=True,
                    help="Use nested (double) quantization for the 4-bit quant constants.")

    p.add_argument("--max_new_tokens", type=int, default=60,
                    help="Maximum number of tokens to generate per example.")
    p.add_argument("--repetition_penalty", type=float, default=1.1,
                    help="Penalty applied to repeated tokens during generation.")
    p.add_argument("--do_sample", action=argparse.BooleanOptionalAction, default=False,
                    help="Use sampling instead of greedy decoding during generation.")

    p.add_argument("--limit", type=int, default=None,
                    help="Only evaluate the first N test examples (useful for a quick smoke test).")
    p.add_argument("--output_excel", type=str, default="flower_evaluation_results_new.xlsx",
                    help="Path to write the Excel evaluation report to.")
    p.add_argument("--thumb_dir", type=str, default="temp_thumbnails",
                    help="Directory to temporarily store image thumbnails for the report.")
    p.add_argument("--keep_thumbnails", action="store_true",
                    help="Don't delete the thumbnail images after building the report.")

    return p.parse_args()


def main():
    args = parse_args()
    set_seed(args.seed)

    os.makedirs(args.thumb_dir, exist_ok=True)

    print("Loading evaluation dataset...")
    _, _, test_subset, label_map = dataset_format(dataset=args.dataset, seed=args.seed)
    if args.limit is not None:
        test_subset = test_subset.select(range(min(args.limit, len(test_subset))))

    print("Loading model and adapter...")
    bnb_config = build_bnb_config(args.load_in_4bit, args.bnb_quant_type, args.bnb_double_quant)
    base_model = AutoModelForImageTextToText.from_pretrained(
        args.base_model_id,
        quantization_config=bnb_config,
        device_map=model_config.device_map,
    )
    model = PeftModel.from_pretrained(base_model, args.adapter_path)
    processor = AutoProcessor.from_pretrained(args.adapter_path)
    model.eval()

    results_data = []

    print("Running inference on test dataset...")
    for idx, example in enumerate(tqdm(test_subset)):
        test_image = example["image"].convert("RGB")

        thumb = test_image.copy()
        thumb.thumbnail((120, 120))
        thumb_path = os.path.join(args.thumb_dir, f"thumb_{idx}.png")
        thumb.save(thumb_path)

        messages = [
            {
                "role": "user",
                "content": [
                    {"type": "image"},
                    {"type": "text", "text": args.prompt},
                ],
            }
        ]

        prompt = processor.apply_chat_template(messages, add_generation_prompt=True)
        inputs = processor(text=prompt, images=[test_image], return_tensors="pt").to(model.device)

        with torch.no_grad():
            generated_ids = model.generate(
                **inputs,
                max_new_tokens=args.max_new_tokens,
                eos_token_id=processor.tokenizer.eos_token_id,
                repetition_penalty=args.repetition_penalty,
                do_sample=args.do_sample,
            )

        generated_text = processor.batch_decode(generated_ids, skip_special_tokens=True)[0]

        actual_label_id = str(example["label"])
        actual_species = label_map.get(actual_label_id, "Unknown")

        predicted_json_str = "{}"
        match_status = "INCORRECT"

        match = re.search(r"\{.*?\}", generated_text, re.DOTALL)
        if match:
            predicted_json_str = match.group(0).replace("\n", "")
            try:
                output_json = json.loads(predicted_json_str)
                if "label_id" in output_json and str(output_json["label_id"]) == actual_label_id:
                    match_status = "CORRECT"
                elif "flower_type" in output_json and str(output_json["flower_type"]).lower() == actual_species.lower():
                    match_status = "CORRECT"
            except Exception:
                pass

        results_data.append({
            "Thumbnail_Path": thumb_path,
            "Actual Label": actual_species,
            "Predicted Label": predicted_json_str,
            "Match Status": match_status,
        })

    print("Writing records out to formatted Excel file...")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Evaluation Report"

    headers = ["Visual Image", "Actual Label", "Predicted Label", "Match Status"]
    ws.append(headers)

    for row_idx, row_data in enumerate(results_data, start=2):
        ws.cell(row=row_idx, column=2, value=row_data["Actual Label"])
        ws.cell(row=row_idx, column=3, value=row_data["Predicted Label"])
        ws.cell(row=row_idx, column=4, value=row_data["Match Status"])

        if os.path.exists(row_data["Thumbnail_Path"]):
            img = OpenPyxlImage(row_data["Thumbnail_Path"])
            ws.add_image(img, f"A{row_idx}")

        ws.row_dimensions[row_idx].height = 95

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 15

    wb.save(args.output_excel)
    print(f"Sheet generated and saved to: {args.output_excel}")

    correct_count = sum(1 for r in results_data if r["Match Status"] == "CORRECT")
    print(f"Accuracy: {correct_count}/{len(results_data)} ({correct_count/len(results_data)*100:.2f}%)")

    if not args.keep_thumbnails:
        for r in results_data:
            if os.path.exists(r["Thumbnail_Path"]):
                os.remove(r["Thumbnail_Path"])
        os.rmdir(args.thumb_dir)

    # Save this eval run's exact args next to the report for reproducibility.
    eval_config_path = os.path.splitext(args.output_excel)[0] + "_eval_config.json"
    with open(eval_config_path, "w") as f:
        json.dump(vars(args), f, indent=2)
    print(f"Eval config saved to {eval_config_path}")


if __name__ == "__main__":
    main()